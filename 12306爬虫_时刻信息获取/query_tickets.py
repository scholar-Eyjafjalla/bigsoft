"""
gt.save_data(trains_data_list, date, begin, end)
第一个参数是城市代码和城市名称的对应文件。由init_station_code生成
调用此函数即可返回出对应的结果，并且生成对应的xlsx文件：
如 gt.save_data( ,2024-11-01,武汉，长沙)
"""

import requests
import json
from openpyxl import Workbook
from prettytable import PrettyTable
from init_station_code import get_city_data


cookies = """
_uab_collina=173176137701442003843301; JSESSIONID=28E7652760A8487389E7D64E16712376; BIGipServerotn=585629962.64545.0000; guidesStatus=off; highContrastMode=defaltMode; cursorStatus=off; BIGipServerpassport=971505930.50215.0000; route=6f50b51faa11b987e576cdb301e545c4; _jc_save_fromStation=%u5317%u4EAC%2CBJP; _jc_save_toStation=%u4E0A%u6D77%2CSHH; _jc_save_fromDate=2024-11-16; _jc_save_toDate=2024-11-16; _jc_save_wfdc_flag=dc
"""

class GetTrains:
    def __init__(self, date, begin_id, end_id):
        # 请求的目标链接
        self.url = "https://kyfw.12306.cn/otn/leftTicket/query"
        # cookies
        self.cookies = {
            '_uab_collina': '171324859263120074949415',
            'JSESSIONID': '708D9C6917F9858184F462E86DC45BD0',
            '_jc_save_fromStation': '%u82CF%u5DDE%2CSZH',
            '_jc_save_toStation': '%u6C5D%u5DDE%2CROF',
            '_jc_save_fromDate': '2024-04-30',
            '_jc_save_wfdc_flag': 'dc',
            'route': '9036359bb8a8a461c164a04f8f50b252',
            'BIGipServerotn': '1172832522.24610.0000',
            'BIGipServerpassport': '854065418.50215.0000',
            'guidesStatus': 'off',
            'highContrastMode': 'defaltMode',
            'cursorStatus': 'off',
            '_jc_save_toDate': '2024-04-30',
        }
        # 构建请求头
        self.headers = {
            'Accept': '*/*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6,zh-TW;q=0.5',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'If-Modified-Since': '0',
            'Pragma': 'no-cache',
            'Referer': 'https://www.12306.cn/index/index.html',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
            'X-Requested-With': 'XMLHttpRequest',
        }
        # 构建请求所需参数
        self.params = {
            "leftTicketDTO.train_date": date,
            "leftTicketDTO.from_station": begin_id,
            "leftTicketDTO.to_station": end_id,
            "purpose_codes": "ADULT"
        }
        # 实例化美化表格对象
        self.pt = PrettyTable()

    def run(self):
        # 对目标网址发送请求
        res = requests.get(
            self.url, headers=self.headers, params=self.params, cookies=self.cookies
        ).json()
        data_list = res['data']['result']
        # 构造表格的表头，用于展示和保存
        header_list = [
            ['车次', '出发时间', '到达时间', '历时', '商务座', '一等座', '二等座', '软卧', '硬卧', '硬座', '无座', '备注']
        ]
        # 将表头信息添加进展示表格的表头
        self.pt.field_names = header_list[0]
        for data in data_list:
            # 格式化添加表数据
            trains_msg = self.format_data(data)
            # 将数据添加进列表，用于保存
            header_list.append(trains_msg)
        # 打印表格
        print(self.pt)
        # 返回车次信息列表
        return header_list

    def format_data(self, data):
        # 将返回的数据以'|'进行分隔
        all_data_list = data.split('|')
        # 提取车次的信息
        trains_msg = [
            all_data_list[3],
            all_data_list[8],
            all_data_list[9],
            all_data_list[10],
            all_data_list[32] if all_data_list[32] != "" else "--",
            all_data_list[31] if all_data_list[31] != "" else "--",
            all_data_list[30] if all_data_list[30] != "" else "--",
            all_data_list[23] if all_data_list[23] != "" else "--",
            all_data_list[28] if all_data_list[28] != "" else "--",
            all_data_list[29] if all_data_list[29] != "" else "--",
            all_data_list[26] if all_data_list[26] != "" else "--",
            all_data_list[1] if all_data_list[1] != "" else "--"
        ]
        # 增添表内容
        self.pt.add_row(trains_msg)
        # 将提取的信息返回，用于保存
        return trains_msg

    def save_data(self, trains_data_list, date, begin, end):
        wb = Workbook()
        sheet = wb.create_sheet("车次信息", -1)
        # 遍历表格索引，写入数据
        for x in range(len(trains_data_list)):
            for y in range(len(trains_data_list[x])):
                sheet.cell(x + 1, y + 1).value = trains_data_list[x][y]
        wb.save(f"{date}_{begin}_{end}.xlsx")
        print("数据保存完成！")


if __name__ == '__main__':
    # 更新城市对应的英文代码，需要时再启用
    # get_city_data()
    date = input("请输入出发日期(YYYY-MM-DD)：")
    begin = input("请输入出发地：")
    end = input("请输入目的地：")
    # 读取生成的json文件
    city_list = json.load(open('city_data.json', 'r'))
    # 获取城市对应的英文代码
    begin_id = city_list[begin]
    end_id = city_list[end]
    gt = GetTrains(date, begin_id, end_id)
    trains_data_list = gt.run()
    # 是否需要保存数据
    gt.save_data(trains_data_list, date, begin, end)
    print(
        "12306直达链接(复制到浏览器打开)：",
        "https://kyfw.12306.cn/otn/leftTicket/init?"
        "linktypeid=dc&"
        f"fs={begin},{begin_id}&"
        f"ts={end},{end_id}&"
        f"date={date}&"
        "flag=N,N,Y"
    )
    print("cookies:",cookies)
